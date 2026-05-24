#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Gergely István — komplett adatkinyerés a dashboardhoz + felfedező elemzések."""
import openpyxl, re, json, os
from collections import defaultdict

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
def P(name): return os.path.join(BASE, name)

D = {}  # dashboard data

# ---------- ADAOS (kategória árrés) ----------
wb = openpyxl.load_workbook(P('Adaos total 2025xlsx.xlsx'), data_only=True, read_only=True)
rows = list(wb.active.iter_rows(values_only=True)); wb.close()
cats = []
for r in rows[2:-1]:
    name, van, tva, adaos, ndoc = r[0], r[1], r[2], r[3], r[4]
    if isinstance(van,(int,float)) and van>0 and isinstance(name,str):
        nm = name.replace('Total','').replace('»','').strip()
        cats.append({
            'kategoria': nm if nm else '(besorolatlan)',
            'forgalom': round(van,2),
            'tva': round(tva or 0,2),
            'adaos': round(adaos or 0,2),
            'arres_pct': round((adaos or 0)/van*100,2),
            'tva_pct': round((tva or 0)/van*100,2),
            'dokumentum': ndoc or 0,
        })
tot = rows[-1]
D['adaos'] = {
    'kategoriak': sorted(cats, key=lambda x:-x['forgalom']),
    'osszesen': {'forgalom': round(tot[1],2), 'tva': round(tot[2],2), 'adaos': round(tot[3],2),
                 'dokumentum': tot[4], 'arres_pct': round(tot[3]/tot[1]*100,2)},
}
# áfa-szerkezet
vat9 = sum(c['forgalom'] for c in cats if c['tva_pct']<14)
vat_std = sum(c['forgalom'] for c in cats if c['tva_pct']>=17)
vat_mid = sum(c['forgalom'] for c in cats if 14<=c['tva_pct']<17)
D['afa_szerkezet'] = {'kedvezmenyes_9': round(vat9,2), 'standard_19_21': round(vat_std,2), 'vegyes': round(vat_mid,2)}

# ---------- ZGY (B2B partnerek) ----------
wb = openpyxl.load_workbook(P('2025ZGYxlsx.xlsx'), data_only=True, read_only=True)
rows = list(wb.active.iter_rows(values_only=True)); wb.close()
partners = []
for r in rows[1:]:
    if isinstance(r[0],str) and not r[0].startswith('T O T A L') and isinstance(r[1],(int,float)) and r[1]>0.01:
        nev = r[0].rsplit(',',1)
        partners.append({'partner': nev[0].strip(), 'telepules': (nev[1].strip() if len(nev)>1 else ''), 'ertek': round(r[1],2)})
partners.sort(key=lambda x:-x['ertek'])
zgy_total = sum(p['ertek'] for p in partners)
# kumulatív + HHI
cum = 0
for p in partners:
    p['reszesedes_pct'] = round(p['ertek']/zgy_total*100,2)
    cum += p['ertek']; p['kumulativ_pct'] = round(cum/zgy_total*100,2)
hhi = sum((p['ertek']/zgy_total*100)**2 for p in partners)
top5 = sum(p['ertek'] for p in partners[:5])/zgy_total*100
top10 = sum(p['ertek'] for p in partners[:10])/zgy_total*100
D['partnerek'] = {'lista': partners, 'osszesen': round(zgy_total,2), 'darab': len(partners),
                  'hhi': round(hhi,1), 'top5_pct': round(top5,1), 'top10_pct': round(top10,1)}

# ---------- P2025 (számlás profit időben) ----------
wb = openpyxl.load_workbook(P('P2025 SZAMLAxlsx.xlsx'), data_only=True, read_only=True)
mcost=defaultdict(float); mval=defaultdict(float); mpr=defaultdict(float); mcnt=defaultdict(int)
gtot_cost=gtot_val=gtot_pr=0; ninv=0
for r in wb.active.iter_rows(values_only=True):
    a=r[0]
    if isinstance(a,str):
        mm=re.match(r'(\d{2})\.(\d{2})\.(\d{4}), curs', a)
        if mm:
            k=mm.group(2)
            if isinstance(r[3],(int,float)): mcost[k]+=r[3]; gtot_cost+=r[3]
            if isinstance(r[4],(int,float)): mval[k]+=r[4]; gtot_val+=r[4]
            if isinstance(r[5],(int,float)): mpr[k]+=r[5]; gtot_pr+=r[5]
            fm=re.search(r'total facturi = (\d+)',a)
            if fm: mcnt[k]+=int(fm.group(1)); ninv+=int(fm.group(1))
wb.close()
D['szamlas_havi'] = [{'honap':k,'ertek':round(mval[k],2),'koltseg':round(mcost[k],2),
                      'profit':round(mpr[k],2),'profit_pct':round(mpr[k]/mval[k]*100,2),'szamlak':mcnt[k]}
                     for k in sorted(mval)]
D['szamlas_ossz'] = {'ertek':round(gtot_val,2),'koltseg':round(gtot_cost,2),'profit':round(gtot_pr,2),
                     'profit_pct':round(gtot_pr/gtot_val*100,2),'szamlak':ninv}

# ---------- GERDIT (számla-regiszter, eladás) ----------
wb = openpyxl.load_workbook(P('TOTAL GERDIT 2025xlsx.xlsx'), data_only=True, read_only=True)
gest=None; gdata=defaultdict(lambda:{'szamlak':0,'fara':0.0,'cu':0.0})
gmon=defaultdict(lambda:defaultdict(float)); allmon=defaultdict(float)
for r in wb.active.iter_rows(values_only=True):
    a=r[0]
    if isinstance(a,str) and a.startswith('Gestiune'):
        gest=a.replace('Gestiune','').strip(); continue
    if isinstance(a,(int,float)) and gest:
        gdata[gest]['szamlak']+=1
        if isinstance(r[2],(int,float)): gdata[gest]['fara']+=r[2]
        if isinstance(r[3],(int,float)): gdata[gest]['cu']+=r[3]
        if isinstance(r[1],str):
            mm=re.search(r'/(\d{2})\.(\d{2})\.(\d{4})',r[1])
            if mm and isinstance(r[2],(int,float)):
                gmon[gest][mm.group(2)]+=r[2]; allmon[mm.group(2)]+=r[2]
wb.close()
gerdit_total=sum(v['fara'] for v in gdata.values())
D['gestiune'] = [{'nev':k,'szamlak':v['szamlak'],'arbevetel':round(v['fara'],2),'arbevetel_afaval':round(v['cu'],2),
                  'atlag_szamla':round(v['fara']/v['szamlak'],2),'reszesedes_pct':round(v['fara']/gerdit_total*100,2)}
                 for k,v in sorted(gdata.items(), key=lambda x:-x[1]['fara'])]
D['gerdit_havi'] = [{'honap':k,'arbevetel':round(allmon[k],2)} for k in sorted(allmon)]
D['gestiune_havi'] = {g:{k:round(gmon[g][k],2) for k in sorted(gmon[g])} for g in gmon}

# ---------- PTOT (készlet) ----------
wb = openpyxl.load_workbook(P('2025 PTOT xlsx.xlsx'), data_only=True, read_only=True)
gest=None
neg=[]; dead=[]; nomove=0; total_rows=0
art_qty=defaultdict(float); art_stores=defaultdict(set)
gest_stock={}
for r in wb.active.iter_rows(values_only=True):
    a=r[0]
    if isinstance(a,str) and a.startswith('Gestiune'):
        gest=a.replace('Gestiune:','').strip()
        gest_stock.setdefault(gest,{'cikk':0,'si':0.0,'intr':0.0,'ie':0.0,'sf':0.0}); continue
    if isinstance(a,(int,float)) and gest:
        total_rows+=1
        art=r[1]; um=r[2]; si=r[3] or 0; intr=r[4] or 0; ie=r[5] or 0; sf=r[6] or 0
        gs=gest_stock[gest]; gs['cikk']+=1; gs['si']+=si; gs['intr']+=intr; gs['ie']+=ie; gs['sf']+=sf
        if isinstance(sf,(int,float)) and sf<0: neg.append({'gestiune':gest,'cikk':art,'um':um,'nyito':si,'be':intr,'ki':ie,'zaro':round(sf,2)})
        if sf and sf>0 and ie==0: dead.append({'gestiune':gest,'cikk':art,'um':um,'be':intr,'zaro':round(sf,2)})
        if intr==0 and ie==0 and si==0 and sf==0: nomove+=1
        if isinstance(art,str): art_qty[art]+=ie if isinstance(ie,(int,float)) else 0; art_stores[art].add(gest)
wb.close()
# forgási mutató gestiune-onként (iesiri / átlagkészlet)
for g,v in gest_stock.items():
    avg=(v['si']+v['sf'])/2
    v['forgas']=round(v['ie']/avg,2) if avg>0 else None
    for kk in ['si','intr','ie','sf']: v[kk]=round(v[kk],1)
D['keszlet_egeszseg'] = {
    'osszes_cikksor': total_rows,
    'negativ_db': len(neg), 'holt_db': len(dead), 'mozdulatlan_db': nomove,
    'gestiune_keszlet': gest_stock,
}
D['top_mozgok'] = [{'cikk':a,'mennyiseg':round(q,0),'egysegek':len(art_stores[a])}
                   for a,q in sorted(art_qty.items(),key=lambda x:-x[1])[:20]]
# szortiment-átfedés
overlap=defaultdict(int)
for a,st in art_stores.items(): overlap[len(st)]+=1
D['szortiment_atfedes'] = {str(k):overlap[k] for k in sorted(overlap)}

# ---------- ÖSSZEGZŐ KPI ----------
D['kpi'] = {
    'teljes_arbevetel': round(gerdit_total,2),
    'kassza_arbevetel': D['adaos']['osszesen']['forgalom'],
    'b2b_arbevetel': D['szamlas_ossz']['ertek'],
    'kassza_pct': round(D['adaos']['osszesen']['forgalom']/gerdit_total*100,1),
    'b2b_pct': round(D['szamlas_ossz']['ertek']/gerdit_total*100,1),
    'kassza_arres': D['adaos']['osszesen']['adaos'],
    'kassza_arres_pct': D['adaos']['osszesen']['arres_pct'],
    'b2b_profit': D['szamlas_ossz']['profit'],
    'b2b_profit_pct': D['szamlas_ossz']['profit_pct'],
    'egyseg_db': len(D['gestiune']),
    'partner_db': D['partnerek']['darab'],
    'rekonciliacio_elteres': round(gerdit_total - (D['adaos']['osszesen']['forgalom']+D['szamlas_ossz']['ertek']),2),
}

# kiírás
os.makedirs(P('_build'), exist_ok=True)
with open(P('_build/dashboard_data.json'),'w',encoding='utf-8') as f:
    json.dump(D,f,ensure_ascii=False,indent=2)
# külön a nehéz listák (negatív/holt) a dashboardba mintaként + teljes az Excelhez
with open(P('_build/keszlet_negativ.json'),'w',encoding='utf-8') as f: json.dump(neg,f,ensure_ascii=False)
with open(P('_build/keszlet_holt.json'),'w',encoding='utf-8') as f: json.dump(dead,f,ensure_ascii=False)

print('=== KPI ===')
for k,v in D['kpi'].items(): print(f'  {k}: {v}')
print('=== Gestiune forgási mutató ===')
for g,v in gest_stock.items(): print(f"  {g:16s} forgas={v['forgas']}  cikk={v['cikk']}")
print('=== Szortiment-átfedés (hány egységben van egy cikk: db cikk) ===', D['szortiment_atfedes'])
print('JSON kiírva: _build/dashboard_data.json')
