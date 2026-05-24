#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Adatréteg v2 — év-kulcsolt, termék-kategorizálással, heti ritmussal.
Kimenet: _build/data_v2.json  ->  { meta:{...}, years:{ "2025": {...} } }
2024 később ugyanígy beilleszthető a years alá; a dashboard év-választóval kezeli."""
import openpyxl, re, json, os, datetime
from collections import defaultdict
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
P = lambda n: os.path.join(BASE, n)

YEAR = "2025"
FILES = {
  'ptot': '2025 PTOT xlsx.xlsx', 'zgy': '2025ZGYxlsx.xlsx', 'adaos': 'Adaos total 2025xlsx.xlsx',
  'p2025': 'P2025 SZAMLAxlsx.xlsx', 'gerdit': 'TOTAL GERDIT 2025xlsx.xlsx',
}

# ---------- META-KATEGÓRIA: "mire költenek" (Adaos kategóriák csoportosítása) ----------
META = {
 'Alkohol & dohány': ['BERE','VIN','BAUT ALCOOLICE','TIGARI'],
 'Üdítő, víz, kávé/tea': ['APA MINERALA','BAUTURI FARA ALCOOL','RACORITOARE','CAFEA','CEAI','SIROP'],
 'Friss élelmiszer': ['LEGUME','LEGUME RO','LEGUME FRUCTE','FRUCTE','OUA','PRODUSE PAINE,PANIFICATIE',
                       'PRODUSE LACTATE','PRODUSE MEZELURI','PRODUSE CONGELATE'],
 'Édesség & snack': ['DULCIURI','BISCUITI','PRAJITURI','CHIPSURI','INGHETATA','PRODUSE SARATE','SEMINTE','GEM'],
 'Alap- és száraz élelmiszer': ['ALTE ALIMENTE','FAINA','OREZ','PASTE FAINOASE','PASTE TOMATE','ULEI','ZAHAR',
                       'CONDIMENTE','CONSERVE','OTET','MURATURA','CEREALE','MARGARINE','PRODUSE DIABETICE','PRODUSE VEG'],
 'Háztartás & vegyiáru': ['DETERGENT','PRODUSE COSMETICE','PRODUSE HARTIE','PRODUSE DIN PLASTIC','PRODUSE INDUSTRIALE'],
 'Non-food egyéb': ['CARTELE TELEFONICE','CIORAPI','RECHIZITE SCOLARE','GRADINA','HRANA ANIMALE'],
 'Egyéb / technikai': ['(besorolatlan)','DISCOUNT','PROMO'],
}
CAT2META = {}
for m, cats in META.items():
    for c in cats: CAT2META[c] = m

# ---------- KULCSSZAVAS TERMÉK-KATEGORIZÁLÓ (cikknév -> Adaos kategória) ----------
# Sorrend számít: az első találat nyer. (kulcsszó-lista, kategória)
RULES = [
 (['BERE'],'BERE'),(['VIN ','VIN','SPUMANT','PROSECCO'],'VIN'),
 (['TIGAR','TIGARI','TUTUN','MARLBORO','WINSTON','KENT','CAMEL','LM ','PALL MALL','ROTHMANS','VICEROY','PARLIAMENT','CHESTERFIELD','DUNHILL'],'TIGARI'),
 (['VODKA','WHISKY','WHISKEY','RUM','ROM ','GIN','LICHIOR','PALINCA','TUICA','VINARS','COGNAC','CONIAC','TEQUILA','HORINCA','BRANDY','APERITIV','VERMUT','UNIREA','JIDVEI'],'BAUT ALCOOLICE'),
 (['APA MIN','APA PLATA','APA TUSNAD','APA BORSEC','APA IZVOR','APA '],'APA MINERALA'),
 (['CAFEA','CAPPUCCINO','NESCAFE','JACOBS','3IN1','3 IN 1','LAVAZZA','TCHIBO'],'CAFEA'),
 (['CEAI'],'CEAI'),(['SIROP'],'SIROP'),
 (['RACORITOARE','COCA','PEPSI','FANTA','SPRITE','ORANGE','SUC ','NECTAR','ICE TEA','LIPTON','CAPPY','PRIGAT','SCHWEPPES','RED BULL','HELL','MONSTER','BURN ','ENERGIZ','MIRINDA','7UP','KINLEY','FUZETEA','TYMBARK'],'RACORITOARE'),
 (['BAUTURA','BAUT '],'BAUTURI FARA ALCOOL'),
 (['PAINE','CHIFLA','FRANZELA','BAGHET','LIPIE','COVRIG','CORN ','PANIF'],'PRODUSE PAINE,PANIFICATIE'),
 (['LAPTE','IAURT','BRANZA','CASCAVAL','SMANTANA','UNT ','TELEMEA','KEFIR','FRISCA','SANA ','URDA','MOZZARELLA'],'PRODUSE LACTATE'),
 (['SALAM','CARNATI','PARIZER','SUNCA','CRENVURSTI','MEZEL','KAIZER','SLANINA','PASTRAMA','TOBA','CABANOS','BACON','JAMBON','PREPARAT'],'PRODUSE MEZELURI'),
 (['OUA','OU '],'OUA'),
 (['CARTOF','ROSII','CEAPA','ARDEI','VARZA','MORCOV','CASTRAVET','USTUROI','DOVLEC','VINETE','RIDICHI','SALATA','SPANAC','PRAZ','TELINA','SFECLA','LEGUME','CONOPIDA','BROCCOLI','FASOLE VERDE','MAZARE'],'LEGUME'),
 (['BANANE','MERE','PORTOCAL','LAMAI','STRUGURI','PRUNE','PERE','PIERSIC','CAISE','CIRESE','CAPSUNI','PEPENE','KIWI','MANDARINE','ANANAS','RODII','AVOCADO','FRUCTE','NUCI','CURMALE','SMOCHINE'],'FRUCTE'),
 (['INGHETATA'],'INGHETATA'),
 (['CIOCOLATA','BOMBOANE','PRALINE','KINDER','NAPOLITAN','BATON','GUMA','JELEU','DROPS','ACADEA','DULCIUR','HALVA','RAHAT','MILKA','KITKAT','SNICKERS','BOUNTY','HARIBO','MENTOS','TIC TAC'],'DULCIURI'),
 (['BISCUIT','EUGENIA','CRACKER','PISCOT'],'BISCUITI'),
 (['PRAJITUR','TORT ','CHEC','COZONAC','GOGOSI','SAVARINA'],'PRAJITURI'),
 (['CHIPS','CHIO','STICKS','POPCORN','PUFULETI','CRUTON'],'CHIPSURI'),
 (['ALUNE','ARAHIDE','SEMINTE','FISTIC','MIGDALE','CAJU'],'SEMINTE'),
 (['STICKS SARAT','COVRIGEI','SARATIN','PRODUSE SARATE','SARATE'],'PRODUSE SARATE'),
 (['DETERGENT','ARIEL','PERSIL','TIDE','DERO','BONUX'],'DETERGENT'),
 (['GEL','CREMA','BALSAM','DEODORANT','SAPUN','SAMPON','PASTA DE DINTI','PERIUTA','SPUMA','LOTIUNE','PARFUM','AFTER SHAVE','MASCARA','RUJ','OJA','VOPSEA PAR','DEO ','ANTIPERSPIRANT','NIVEA','DOVE','REXONA','GILLETTE','COLGATE','HEAD','SCHWARZKOPF','LOREAL','ABSORBANT','TAMPOANE','SCUTEC','PAMPERS','SERVETELE UMEDE'],'PRODUSE COSMETICE'),
 (['HARTIE','SERVETELE','PROSOP HARTIE','HARTIE IGIENICA','SERVET'],'PRODUSE HARTIE'),
 (['PUNGI','SACI','FOLIE','PLASTIC','PAHARE PLASTIC','TACAMURI PLASTIC','CALD','PET '],'PRODUSE DIN PLASTIC'),
 (['ODORIZANT','SOLUTIE','BURETE','MANUSI','LAVETA','MATURA','GALEATA','CLOR','ACID','ABURBA','APARAT WC','ODORIZ','RAID','PRONTO','CIF ','DOMESTOS','BATERII','BEC ','BRICHET','LUMANARE','CHIBRIT','BATERIE','SET '],'PRODUSE INDUSTRIALE'),
 (['FAINA'],'FAINA'),(['OREZ'],'OREZ'),(['ZAHAR'],'ZAHAR'),
 (['PASTE','TAITEI','SPAGHETE','MACAROANE','FIDEA','FUSILLI','PENNE'],'PASTE FAINOASE'),
 (['BULION','PASTA DE TOMATE','PASTA TOMATE','SUC ROSII'],'PASTE TOMATE'),
 (['ULEI','UNTDELEMN'],'ULEI'),(['OTET'],'OTET'),(['MARGARINA','MARGARINE','RAMA ','UNT VEGETAL'],'MARGARINE'),
 (['SARE','PIPER','CONDIMENT','BORS','VEGETA','DELIKAT','BOIA','CIMBRU','SCORTISOARA','VANILIE','PRAF DE COPT','DROJDIE','GELATINA','BICARBONAT'],'CONDIMENTE'),
 (['CONSERVA','PATEU','PESTE CONSERVA','TON ','MACROU','SARDINE','FASOLE BOABE','MAZARE CONSERVA','PORUMB CONSERVA','COMPOT'],'CONSERVE'),
 (['MURATUR','CASTRAVETI MURATI','GOGOSARI','ZACUSCA','ARDEI IUTE MURAT'],'MURATURA'),
 (['CEREALE','FULGI','MUSLI','CORN FLAKES','NESQUIK'],'CEREALE'),
 (['GEM','DULCEATA','MAGIUN','MIERE'],'GEM'),
 (['CONGELAT','GHEATA','PIZZA CONG','LEGUME CONG'],'PRODUSE CONGELATE'),
 (['HRANA','FELIX','PEDIGREE','WHISKAS','PURINA','FRISKIES','CATEI','PISICI'],'HRANA ANIMALE'),
 (['CARTELA','RECARG','VOUCHER','ORANGE PREPAY','VODAFONE','TELEKOM','DIGI'],'CARTELE TELEFONICE'),
 (['SOSETE','CIORAPI','DRESURI'],'CIORAPI'),
 (['CAIET','PIX','CREION','RECHIZIT','PENAR','GUMA STERS','MARKER','LIPICI','RIGLA','GHIOZDAN'],'RECHIZITE SCOLARE'),
 (['SEMINTE GRADINA','RASAD','INGRASAMANT','GRADINA','GHIVECI'],'GRADINA'),
 (['DIABETIC'],'PRODUSE DIABETICE'),
 (['CHIBRITURI'],'PRODUSE INDUSTRIALE'),
]
def classify(name):
    u = name.upper()
    for kws, cat in RULES:
        for kw in kws:
            if kw in u: return cat
    return '(besorolatlan)'

YD = {}  # year data

# ---------- ADAOS (kategória, érték = mire költenek) ----------
wb = openpyxl.load_workbook(P(FILES['adaos']), data_only=True, read_only=True)
rows = list(wb.active.iter_rows(values_only=True)); wb.close()
cats = []
for r in rows[2:-1]:
    name, van, tva, adaos, ndoc = r[0], r[1], r[2], r[3], r[4]
    if isinstance(van,(int,float)) and van>0 and isinstance(name,str):
        nm = name.replace('Total','').replace('»','').strip() or '(besorolatlan)'
        cats.append({'kategoria':nm,'meta':CAT2META.get(nm,'Egyéb / technikai'),
                     'forgalom':round(van,2),'tva':round(tva or 0,2),'adaos':round(adaos or 0,2),
                     'arres_pct':round((adaos or 0)/van*100,2),'tva_pct':round((tva or 0)/van*100,2),'dokumentum':ndoc or 0})
tot = rows[-1]; TF = tot[1]
for c in cats: c['forgalom_pct'] = round(c['forgalom']/TF*100,2)
# meta aggregálás
meta_agg = defaultdict(lambda:{'forgalom':0.0,'adaos':0.0,'tva':0.0,'kategoriak':[]})
for c in cats:
    m = meta_agg[c['meta']]; m['forgalom']+=c['forgalom']; m['adaos']+=c['adaos']; m['tva']+=c['tva']; m['kategoriak'].append(c['kategoria'])
meta_list = [{'meta':k,'forgalom':round(v['forgalom'],2),'forgalom_pct':round(v['forgalom']/TF*100,2),
              'adaos':round(v['adaos'],2),'arres_pct':round(v['adaos']/v['forgalom']*100,2) if v['forgalom'] else 0,
              'kategoriak':v['kategoriak']} for k,v in meta_agg.items()]
meta_list.sort(key=lambda x:-x['forgalom'])
YD['adaos'] = {'kategoriak':sorted(cats,key=lambda x:-x['forgalom']),'meta':meta_list,
   'osszesen':{'forgalom':round(tot[1],2),'tva':round(tot[2],2),'adaos':round(tot[3],2),'dokumentum':tot[4],'arres_pct':round(tot[3]/tot[1]*100,2)}}
YD['afa_szerkezet']={'kedvezmenyes_9':round(sum(c['forgalom'] for c in cats if c['tva_pct']<14),2),
   'standard_19_21':round(sum(c['forgalom'] for c in cats if c['tva_pct']>=17),2),
   'vegyes':round(sum(c['forgalom'] for c in cats if 14<=c['tva_pct']<17),2)}

# ---------- ZGY partnerek ----------
wb = openpyxl.load_workbook(P(FILES['zgy']), data_only=True, read_only=True)
rows=list(wb.active.iter_rows(values_only=True)); wb.close()
partners=[]
for r in rows[1:]:
    if isinstance(r[0],str) and not r[0].startswith('T O T A L') and isinstance(r[1],(int,float)) and r[1]>0.01:
        nv=r[0].rsplit(',',1); partners.append({'partner':nv[0].strip(),'telepules':(nv[1].strip() if len(nv)>1 else ''),'ertek':round(r[1],2)})
partners.sort(key=lambda x:-x['ertek']); ZT=sum(p['ertek'] for p in partners); cum=0
for p in partners:
    p['reszesedes_pct']=round(p['ertek']/ZT*100,2); cum+=p['ertek']; p['kumulativ_pct']=round(cum/ZT*100,2)
YD['partnerek']={'lista':partners,'osszesen':round(ZT,2),'darab':len(partners),
   'hhi':round(sum((p['ertek']/ZT*100)**2 for p in partners),1),
   'top5_pct':round(sum(p['ertek'] for p in partners[:5])/ZT*100,1),
   'top10_pct':round(sum(p['ertek'] for p in partners[:10])/ZT*100,1)}

# ---------- P2025 (B2B idő, profit, heti ritmus) ----------
wb=openpyxl.load_workbook(P(FILES['p2025']),data_only=True,read_only=True)
mcost=defaultdict(float);mval=defaultdict(float);mpr=defaultdict(float);mcnt=defaultdict(int)
dow=defaultdict(lambda:[0.0,0.0,0]); ninv=0; gc=gv=gp=0
for r in wb.active.iter_rows(values_only=True):
    a=r[0]
    if isinstance(a,str):
        mm=re.match(r'(\d{2})\.(\d{2})\.(\d{4}), curs',a)
        if mm:
            k=mm.group(2)
            if isinstance(r[3],(int,float)):mcost[k]+=r[3];gc+=r[3]
            if isinstance(r[4],(int,float)):mval[k]+=r[4];gv+=r[4]
            if isinstance(r[5],(int,float)):mpr[k]+=r[5];gp+=r[5]
            fm=re.search(r'total facturi = (\d+)',a)
            if fm:mcnt[k]+=int(fm.group(1));ninv+=int(fm.group(1))
            if isinstance(r[4],(int,float)):
                d=datetime.date(int(mm.group(3)),int(mm.group(2)),int(mm.group(1)))
                dw=dow[d.weekday()];dw[0]+=r[4];dw[1]+=(r[5] or 0);dw[2]+=1
wb.close()
MONTHS=[f'{i:02d}' for i in range(1,13)]
YD['szamlas_havi']=[{'honap':k,'ertek':round(mval[k],2),'koltseg':round(mcost[k],2),'profit':round(mpr[k],2),
   'profit_pct':round(mpr[k]/mval[k]*100,2) if mval[k] else 0,'szamlak':mcnt[k]} for k in MONTHS]
YD['szamlas_ossz']={'ertek':round(gv,2),'koltseg':round(gc,2),'profit':round(gp,2),'profit_pct':round(gp/gv*100,2),'szamlak':ninv}
hu=['Hétfő','Kedd','Szerda','Csütörtök','Péntek','Szombat','Vasárnap']
YD['heti_ritmus']=[{'nap':hu[wd],'napok':dow[wd][2],'ertek':round(dow[wd][0],2),
   'ertek_pct':round(dow[wd][0]/gv*100,2),'atlag_nap':round(dow[wd][0]/dow[wd][2],2) if dow[wd][2] else 0,
   'profit_pct':round(dow[wd][1]/dow[wd][0]*100,2) if dow[wd][0] else 0} for wd in range(7)]

# ---------- GERDIT (árbevétel/telephely + havi/telephely) ----------
wb=openpyxl.load_workbook(P(FILES['gerdit']),data_only=True,read_only=True)
gest=None; gdata=defaultdict(lambda:{'szamlak':0,'fara':0.0,'cu':0.0}); gmon=defaultdict(lambda:defaultdict(float)); allmon=defaultdict(float)
for r in wb.active.iter_rows(values_only=True):
    a=r[0]
    if isinstance(a,str) and a.startswith('Gestiune'): gest=a.replace('Gestiune','').strip(); continue
    if isinstance(a,(int,float)) and gest:
        gdata[gest]['szamlak']+=1
        if isinstance(r[2],(int,float)): gdata[gest]['fara']+=r[2]
        if isinstance(r[3],(int,float)): gdata[gest]['cu']+=r[3]
        if isinstance(r[1],str):
            mm=re.search(r'/(\d{2})\.(\d{2})\.(\d{4})',r[1])
            if mm and isinstance(r[2],(int,float)): gmon[gest][mm.group(2)]+=r[2]; allmon[mm.group(2)]+=r[2]
wb.close()
GT=sum(v['fara'] for v in gdata.values())
YD['gestiune']=[{'nev':k,'szamlak':v['szamlak'],'arbevetel':round(v['fara'],2),'arbevetel_afaval':round(v['cu'],2),
   'atlag_szamla':round(v['fara']/v['szamlak'],2),'reszesedes_pct':round(v['fara']/GT*100,2)}
   for k,v in sorted(gdata.items(),key=lambda x:-x[1]['fara'])]
YD['gerdit_havi']=[{'honap':k,'arbevetel':round(allmon.get(k,0),2)} for k in MONTHS]
YD['gestiune_havi']={g:{k:round(gmon[g].get(k,0),2) for k in MONTHS} for g in gmon}

# ---------- PTOT (készlet + termék-kategorizálás) ----------
wb=openpyxl.load_workbook(P(FILES['ptot']),data_only=True,read_only=True)
gest=None; total_rows=0
neg=[]; dead=[]; nomove=0
gest_stock=defaultdict(lambda:{'cikk':0,'si':0.0,'intr':0.0,'ie':0.0,'sf':0.0})
prod=defaultdict(lambda:{'um':'','cat':None,'ie':0.0,'stores':{}})
matched=0
for r in wb.active.iter_rows(values_only=True):
    a=r[0]
    if isinstance(a,str) and a.startswith('Gestiune'): gest=a.replace('Gestiune:','').strip(); continue
    if isinstance(a,(int,float)) and gest and isinstance(r[1],str):
        total_rows+=1
        art=r[1]; um=r[2] or ''; si=r[3] or 0; intr=r[4] or 0; ie=r[5] or 0; sf=r[6] or 0
        gs=gest_stock[gest]; gs['cikk']+=1; gs['si']+=si; gs['intr']+=intr; gs['ie']+=ie; gs['sf']+=sf
        cat=classify(art);
        if cat!='(besorolatlan)': matched+=1
        if isinstance(sf,(int,float)) and sf<0: neg.append({'g':gest,'n':art,'um':um,'si':round(si,2),'in':round(intr,2),'out':round(ie,2),'sf':round(sf,2),'cat':cat})
        if sf and sf>0 and ie==0: dead.append({'g':gest,'n':art,'um':um,'in':round(intr,2),'sf':round(sf,2),'cat':cat})
        if intr==0 and ie==0 and si==0 and sf==0: nomove+=1
        p=prod[art]; p['um']=um; p['cat']=cat; p['ie']+=ie if isinstance(ie,(int,float)) else 0
        p['stores'][gest]={'si':round(si,2),'in':round(intr,2),'out':round(ie,2),'sf':round(sf,2)}
wb.close()
for g,v in gest_stock.items():
    avg=(v['si']+v['sf'])/2; v['forgas']=round(v['ie']/avg,2) if avg>0 else None
    for kk in ['si','intr','ie','sf']: v[kk]=round(v[kk],1)
# termék-kategória mennyiségi aggregálás (cikkszám + össz kiadás kategóriánként és metánként)
cat_q=defaultdict(lambda:{'cikk':0,'ie':0.0}); meta_q=defaultdict(lambda:{'cikk':0,'ie':0.0})
products=[]
for name,p in prod.items():
    cat_q[p['cat']]['cikk']+=1; cat_q[p['cat']]['ie']+=p['ie']
    mm=CAT2META.get(p['cat'],'Egyéb / technikai'); meta_q[mm]['cikk']+=1; meta_q[mm]['ie']+=p['ie']
    products.append({'n':name,'um':p['um'],'cat':p['cat'],'meta':mm,'ie':round(p['ie'],1),
                     'eg':len(p['stores']),'st':p['stores']})
products.sort(key=lambda x:-x['ie'])
YD['keszlet_egeszseg']={'osszes_cikksor':total_rows,'negativ_db':len(neg),'holt_db':len(dead),
   'mozdulatlan_db':nomove,'gestiune_keszlet':dict(gest_stock),
   'kategorizalas_lefedettseg_pct':round(matched/total_rows*100,1)}
YD['keszlet_negativ']=neg
YD['keszlet_holt']=sorted(dead,key=lambda x:-x['sf'])
YD['termekek']=products
YD['termek_kategoria_mennyiseg']={'kategoria':{k:{'cikk':v['cikk'],'ie':round(v['ie'],0)} for k,v in cat_q.items()},
   'meta':{k:{'cikk':v['cikk'],'ie':round(v['ie'],0)} for k,v in meta_q.items()}}
ov=defaultdict(int)
for p in products: ov[p['eg']]+=1
YD['szortiment_atfedes']={str(k):ov[k] for k in sorted(ov)}

# ---------- KPI (csak TÉNYEK, semmi becslés) ----------
YD['kpi']={'teljes_arbevetel':round(GT,2),'kassza_arbevetel':YD['adaos']['osszesen']['forgalom'],
   'b2b_arbevetel':YD['szamlas_ossz']['ertek'],
   'kassza_pct':round(YD['adaos']['osszesen']['forgalom']/GT*100,1),
   'b2b_pct':round(YD['szamlas_ossz']['ertek']/GT*100,1),
   'kassza_arres':YD['adaos']['osszesen']['adaos'],'kassza_arres_pct':YD['adaos']['osszesen']['arres_pct'],
   'b2b_profit':YD['szamlas_ossz']['profit'],'b2b_profit_pct':YD['szamlas_ossz']['profit_pct'],
   'egyseg_db':len(YD['gestiune']),'partner_db':YD['partnerek']['darab'],
   'rekonciliacio_elteres':round(GT-(YD['adaos']['osszesen']['forgalom']+YD['szamlas_ossz']['ertek']),2)}

OUT={'meta':{'generalt':datetime.date.today().isoformat(),'penznem':'RON (lej)','aktiv_ev':YEAR,
     'elerheto_evek':[YEAR],'meta_kategoriak':list(META.keys()),
     'megjegyzes':'A telephelyi bontás csak árbevételre (GERDIT), készletre (PTOT) és havi trendre érhető el. '
                  'A kategória/partner/B2B adat HÁLÓZATI szintű. Termék-kategória: kulcsszavas automatikus besorolás.'},
     'years':{YEAR:YD}}
os.makedirs(P('_build'),exist_ok=True)
json.dump(OUT,open(P('_build/data_v2.json'),'w',encoding='utf-8'),ensure_ascii=False)
print('Kategorizálás lefedettség:',YD['keszlet_egeszseg']['kategorizalas_lefedettseg_pct'],'%')
print('Termékek (distinct):',len(products),'| Negatív:',len(neg),'| Holt:',len(dead))
print('=== Meta-kategória (mire költenek, érték) ===')
for m in meta_list: print(f"  {m['meta']:28s} {m['forgalom']:11.0f} lej  ({m['forgalom_pct']:4.1f}%)  árrés {m['arres_pct']:.1f}%")
print('JSON:', P('_build/data_v2.json'), round(os.path.getsize(P('_build/data_v2.json'))/1024),'KB')
