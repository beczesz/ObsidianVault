#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Készlet-problémák Excel a tulajdonosnak."""
import json, os, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
def P(n): return os.path.join(BASE,n)
neg = json.load(open(P('_build/keszlet_negativ.json'),encoding='utf-8'))
dead = json.load(open(P('_build/keszlet_holt.json'),encoding='utf-8'))
D = json.load(open(P('_build/dashboard_data.json'),encoding='utf-8'))

wb = openpyxl.Workbook()
HEAD = PatternFill('solid', fgColor='1F4E78'); HF = Font(color='FFFFFF', bold=True)
TITLE = Font(size=14, bold=True, color='1F4E78')
thin = Border(*[Side(style='thin', color='D9D9D9')]*4)
RED = PatternFill('solid', fgColor='FCE4E4'); YEL = PatternFill('solid', fgColor='FFF2CC')

def sheet(ws, title, headers, data, fills=None):
    ws['A1']=title; ws['A1'].font=TITLE; ws.append([])
    hr=3
    for c,h in enumerate(headers,1):
        cell=ws.cell(hr,c,h); cell.fill=HEAD; cell.font=HF; cell.alignment=Alignment(horizontal='center')
    for row in data:
        ws.append(row)
    # widths + borders
    for c in range(1,len(headers)+1):
        ws.column_dimensions[get_column_letter(c)].width = 16 if c>2 else (42 if c==2 else 18)
        for r in range(hr, ws.max_row+1):
            ws.cell(r,c).border=thin
    ws.freeze_panes=f'A{hr+1}'
    if fills:
        for r in range(hr+1, ws.max_row+1):
            for c in range(1,len(headers)+1): ws.cell(r,c).fill=fills

# 0 — Összefoglaló
ws=wb.active; ws.title='Osszefoglalo'
ws['A1']='Készlet-problémák — 2025 (forrás: PTOT)'; ws['A1'].font=TITLE
rows=[['',''],
      ['Megnevezés','Darab'],
      ['Összes vizsgált cikksor', D['keszlet_egeszseg']['osszes_cikksor']],
      ['⚠️ Negatív zárókészlet (hiba)', D['keszlet_egeszseg']['negativ_db']],
      ['🟡 Holt készlet (0 éves eladás, van záró)', D['keszlet_egeszseg']['holt_db']],
      ['⚪ Teljesen mozdulatlan (törzsadat-gyanú)', D['keszlet_egeszseg']['mozdulatlan_db']],
      ['',''],
      ['Teendő','']]
for r in rows: ws.append(r)
ws.append(['1) Negatív készlet → leltár/könyvelés kivizsgálás (fizikailag lehetetlen)'])
ws.append(['2) Holt készlet → kifuttatás / akció (lekötött forgótőke)'])
ws.append(['3) Mozdulatlan → cikktörzs tisztítása (duplikált/megszűnt)'])
ws.column_dimensions['A'].width=52; ws.column_dimensions['B'].width=14
for r in range(3,7): ws.cell(r,1).font=Font(bold=True) if r==3 else Font()
ws['A4'].fill=HEAD; ws['B4'].fill=HEAD; ws['A4'].font=HF; ws['B4'].font=HF

# 1 — Negatív
ws=wb.create_sheet('Negativ keszlet')
sheet(ws,'Negatív zárókészlet — kivizsgálandó',
      ['Gestiune','Cikk','UM','Nyitó','Be','Ki','Záró'],
      [[d['gestiune'],d['cikk'],d['um'],d['nyito'],d['be'],d['ki'],d['zaro']] for d in neg], RED)
# 2 — Holt
ws=wb.create_sheet('Holt keszlet')
sheet(ws,'Holt készlet — 0 éves eladás, van zárókészlet',
      ['Gestiune','Cikk','UM','Bevét (év)','Zárókészlet'],
      [[d['gestiune'],d['cikk'],d['um'],d['be'],d['zaro']] for d in sorted(dead,key=lambda x:-x['zaro'])], YEL)

out=P('Keszlet_problemak_2025.xlsx')
wb.save(out)
print('Mentve:', out)
print('Negatív:', len(neg), '| Holt:', len(dead))
